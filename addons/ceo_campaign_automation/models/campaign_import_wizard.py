# -*- coding: utf-8 -*-
import io
import csv
import json
import base64
import requests
import re
import unicodedata
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class CampaignImportWizard(models.TransientModel):
    _name = 'campaign.import.wizard'
    _description = 'Assistant d\'importation et de mapping de contacts IA'

    file_data = fields.Binary(string="Fichier Contacts (CSV ou Excel)")
    file_name = fields.Char(string="Nom du fichier")
    source_tag = fields.Selection([
        ('excel', 'Import Excel/CSV'),
        ('maps', 'Scraping Google Maps'),
        ('annuaire', 'Scraping Annuaire'),
        ('form', 'Formulaire / Landing Page'),
        ('rs', 'Réseaux Sociaux')
    ], string="Source des contacts", default='excel', required=True)

    summary = fields.Text(string="Résumé de l'importation", readonly=True)
    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('review', 'Révision'),
        ('done', 'Terminé')
    ], default='draft')

    # UX fields for review/confirmation
    column_line_ids = fields.One2many(
        'campaign.import.wizard.line', 'wizard_id', string="Colonnes détectées",
    )
    row_limit = fields.Integer(
        string="Nombre de lignes à importer (0 = toutes)",
        default=0,
        help="Laissez à 0 pour importer toutes les lignes. Sinon, seules les N "
             "premières lignes du fichier seront importées.",
    )
    total_rows = fields.Integer(string="Lignes détectées dans le fichier", readonly=True)
    new_fields_count = fields.Integer(
        compute='_compute_counts', string="Nouveaux champs proposés",
    )
    selected_columns_count = fields.Integer(
        compute='_compute_counts', string="Colonnes sélectionnées",
    )

    @api.depends('column_line_ids.is_new', 'column_line_ids.create_field',
                 'column_line_ids.selected')
    def _compute_counts(self):
        for wiz in self:
            wiz.new_fields_count = len(wiz.column_line_ids.filtered(
                lambda l: l.is_new and l.create_field and l.selected))
            wiz.selected_columns_count = len(wiz.column_line_ids.filtered('selected'))

    # ------------------------------------------------------------------ #
    # Helpers
    # ------------------------------------------------------------------ #
    def _sanitize_field_name(self, name):
        """Sanitize a column name into a valid Odoo field name (prefixed by x_)."""
        name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
        name = name.lower()
        name = re.sub(r'[^a-z0-9_]', '_', name)
        name = re.sub(r'_+', '_', name)
        name = name.strip('_')
        if not name.startswith('x_'):
            name = 'x_' + name
        if len(name) > 50:
            name = name[:50].rstrip('_')
        return name

    def _call_gemini_to_map(self, header, sample_values, partner_fields, api_key):
        """Call Google Gemini to map/suggest a target field for a given column."""
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
        headers = {'Content-Type': 'application/json'}

        fields_desc = []
        for name, info in partner_fields.items():
            fields_desc.append(f"- '{name}' (Description: {info.get('string')}, Type: {info.get('type')})")
        fields_str = "\n".join(fields_desc)

        prompt = f"""
Tu es un administrateur de base de données Odoo expert.
Nous importons un fichier de contacts dans le modèle Odoo 'res.partner'.
Nous avons une colonne non mappée avec l'en-tête suivant : "{header}"
Voici les premières valeurs d'exemple de cette colonne :
{json.dumps(sample_values)}

Champs Odoo 'res.partner' existants disponibles pour le mapping :
{fields_str}

Tâches :
1. Détermine si cette colonne correspond à l'un des champs Odoo EXISTANTS. Si oui, mappe-la.
2. Si elle ne correspond pas de façon évidente, suggère la création d'un NOUVEAU champ personnalisé.
   - Le nom technique suggéré doit commencer par 'x_' (ex: 'x_facebook_url', 'x_wilaya_code').
   - Garde le nom en minuscules, alphanumérique, sans espaces ni caractères spéciaux, et sous 40 caractères.
   - Suggère un libellé clair et professionnel en français.

Réponds UNIQUEMENT avec un objet JSON valide. Pour un champ existant :
{{"type": "existing", "field_name": "nom_du_champ_existant"}}
Pour un nouveau champ :
{{"type": "new", "suggested_field_name": "x_nom_suggere", "suggested_label": "Libellé"}}
"""
        payload = {"contents": [{"parts": [{"text": prompt}]}]}

        try:
            response = requests.post(url, headers=headers, json=payload, timeout=15)
            if response.status_code == 200:
                res_json = response.json()
                text_response = res_json.get('candidates', [{}])[0].get(
                    'content', {}).get('parts', [{}])[0].get('text', '').strip()
                if text_response.startswith('```'):
                    lines = text_response.split('\n')
                    if lines[0].startswith('```'):
                        text_response = '\n'.join(lines[1:-1])
                return json.loads(text_response.strip())
        except Exception:
            pass
        return None

    def _detect_field_mapping(self, header, sample_values, partner_fields, gemini_api_key):
        """Detect target field for a column. Returns (field_name, label, is_new)."""
        header_lower = header.lower().strip()

        # 1. Exact match with template standard columns
        from .contact_template_wizard import STANDARD_COLUMNS
        for col in STANDARD_COLUMNS:
            if header_lower in (col['label'].lower(), col['field'].lower()):
                if col['field'] in partner_fields:
                    return col['field'], partner_fields[col['field']]['string'], False

        # 2. Exact match with any existing field's string or name
        for fname, finfo in partner_fields.items():
            if header_lower in (fname.lower(), str(finfo.get('string', '')).lower()):
                return fname, finfo.get('string', ''), False

        clean_header = "".join([c for c in header_lower if c.isalnum() or c.isspace()]).strip()

        mapping_patterns = {
            'name': ['name', 'nom', 'fullname', 'full name', 'raison sociale', 'nom complet',
                     'entreprise', 'company', 'societe', 'société', 'client', 'contact'],
            'email': ['email', 'e-mail', 'mail', 'courriel', 'email address',
                      'adresse email', 'adresse mail'],
            'phone': ['phone', 'téléphone', 'tel', 'telephone', 'tél',
                      'phone number', 'fixe', 'numéro fixe'],
            'mobile': ['mobile', 'portable', 'gsm', 'cellphone', 'cell',
                       'numéro portable', 'téléphone portable'],
            'street': ['street', 'adresse', 'rue', 'address', 'localisation'],
            'city': ['city', 'ville', 'region', 'wilaya', 'commune'],
            'zip': ['zip', 'code postal', 'zipcode', 'cp'],
            'website': ['website', 'site web', 'site', 'url', 'lien'],
            'function': ['function', 'job title', 'poste', 'fonction', 'activité',
                         'domaine', 'secteur', 'role', 'rôle'],
            'comment': ['comment', 'notes', 'description', 'preferences',
                        'commentaire', 'remarque'],
            'vat': ['vat', 'nif', 'tva', 'rc', 'nis', 'registre du commerce'],
        }
        for field, keywords in mapping_patterns.items():
            if clean_header in keywords:
                if field in partner_fields:
                    return field, partner_fields[field]['string'], False

        # Heuristic on values
        email_regex = re.compile(r'[^@\s]+@[^@\s]+\.[^@\s]+')
        phone_regex = re.compile(r'^\+?[0-9\s\-\.\(\)]{8,20}$')
        url_regex = re.compile(r'^(https?://)?(www\.)?[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,}(\/\S*)?$')
        email_votes = phone_votes = url_votes = total_valid = 0
        for val in sample_values:
            if not val:
                continue
            total_valid += 1
            if email_regex.match(val):
                email_votes += 1
            if phone_regex.match(val):
                phone_votes += 1
            if url_regex.match(val):
                url_votes += 1
        if total_valid > 0:
            if email_votes / total_valid > 0.6 and 'email' in partner_fields:
                return 'email', partner_fields['email']['string'], False
            if phone_votes / total_valid > 0.6:
                if 'mobile' in partner_fields:
                    return 'mobile', partner_fields['mobile']['string'], False
                elif 'phone' in partner_fields:
                    return 'phone', partner_fields['phone']['string'], False
            if url_votes / total_valid > 0.6 and 'website' in partner_fields:
                return 'website', partner_fields['website']['string'], False

        # AI fallback
        if gemini_api_key:
            mapping = self._call_gemini_to_map(header, sample_values, partner_fields, gemini_api_key)
            if mapping:
                map_type = mapping.get('type')
                if map_type == 'existing' and mapping.get('field_name') in partner_fields:
                    fname = mapping.get('field_name')
                    return fname, partner_fields[fname]['string'], False
                elif map_type == 'new' and mapping.get('suggested_field_name'):
                    fname = self._sanitize_field_name(mapping.get('suggested_field_name'))
                    flabel = mapping.get('suggested_label') or header
                    return fname, flabel, True

        # Default: propose a new field
        return self._sanitize_field_name(header), header, True

    def _create_custom_field(self, field_name, label, help_text=None):
        """Create a manual field on res.partner and add it to the inherited view."""
        partner_model = self.env['ir.model']._get('res.partner')
        existing = self.env['ir.model.fields'].search([
            ('model_id', '=', partner_model.id),
            ('name', '=', field_name),
        ])
        if not existing:
            self.env['ir.model.fields'].create({
                'name': field_name,
                'model_id': partner_model.id,
                'field_description': label,
                'help': help_text or '',
                'ttype': 'char',
                'state': 'manual',
            })
            self._update_partner_form_view(field_name)
            # Odoo 19: setup_models() was removed from the public Registry API.
            # clear_cache() + invalidate_all() is the correct replacement to
            # flush ORM caches and make the new manual field available in
            # subsequent ORM calls within the same transaction.
            self.env.registry.clear_cache()
            self.env.invalidate_all()
        return field_name

    def _update_partner_form_view(self, field_name=None):
        """
        Inject ALL existing manual x_* fields into a single sibling dynamic
        view that inherits directly from base.view_partner_form (same parent
        as the static CEO tab view).  This avoids the triple-inheritance chain
        that Odoo 19 rejects, and overwrites the arch each time so there are
        never duplicate <field/> tags.
        """
        View = self.env['ir.ui.view']

        # Collect every manual x_* field on res.partner
        partner_model = self.env['ir.model']._get('res.partner')
        custom_fields = self.env['ir.model.fields'].search([
            ('model_id', '=', partner_model.id),
            ('name', '=like', 'x_%'),
            ('state', '=', 'manual'),
        ], order='name')
        if not custom_fields:
            return

        # Build ONE xpath that replaces the entire group content each time
        # The target page/group already exist in the merged arch thanks to the
        # static view (view_partner_form_dynamic_fields), which runs first at
        # its default priority.  Our sibling view runs at priority=99 (after).
        fields_xml = '\n                '.join(
            f'<field name="{cf.name}" string="{cf.field_description}"/>'
            for cf in custom_fields
        )
        arch = (
            "<data>"
            "<xpath expr=\"//page[@name='ai_imported_fields']"
            "/group[@name='dynamic_fields_group']\" position=\"replace\">\n"
            "            <group name=\"dynamic_fields_group\" "
            "string=\"Champs personnalisés\" col=\"4\">\n"
            f"                {fields_xml}\n"
            "            </group>\n"
            "        </xpath></data>"
        )

        # The base form view is the root ancestor — inherit from it directly
        base_form = self.env.ref('base.view_partner_form', raise_if_not_found=False)
        if not base_form:
            return

        DYNAMIC_VIEW_NAME = 'res.partner.ceo.dynamic.fields'
        dynamic_view = View.search([('name', '=', DYNAMIC_VIEW_NAME)], limit=1)
        if dynamic_view:
            # Overwrite — idempotent, no duplicates
            dynamic_view.write({'arch': arch})
        else:
            View.create({
                'name': DYNAMIC_VIEW_NAME,
                'type': 'form',
                'model': 'res.partner',
                'inherit_id': base_form.id,
                'priority': 99,   # run after the static CEO tab view
                'arch': arch,
            })

        self.env.registry.clear_cache()
        self.env.invalidate_all()

    # ------------------------------------------------------------------ #
    # File parsing
    # ------------------------------------------------------------------ #
    def _parse_uploaded_file(self):
        """Decode the uploaded file and return (headers, list[dict])."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier à importer."))
        file_content = base64.b64decode(self.file_data)
        file_name = self.file_name or ""
        is_excel = file_name.lower().endswith(('.xlsx', '.xls'))
        headers, rows = [], []

        if is_excel:
            if not openpyxl:
                raise UserError(_(
                    "Le module 'openpyxl' n'est pas disponible. "
                    "Veuillez importer au format CSV."))
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = wb.active
                header_row = next(sheet.iter_rows(values_only=True), None)
                if not header_row:
                    raise UserError(_("Le fichier Excel est vide."))
                headers = [str(h).strip() for h in header_row if h is not None]
                for r in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(val is not None for val in r):
                        continue
                    row_dict = {}
                    for idx, val in enumerate(r):
                        if idx < len(headers):
                            row_dict[headers[idx]] = str(val).strip() if val is not None else ""
                    rows.append(row_dict)
            except UserError:
                raise
            except Exception as e:
                raise UserError(_("Erreur lors de la lecture du fichier Excel : %s") % str(e))
        else:
            try:
                try:
                    csv_text = file_content.decode('utf-8')
                except UnicodeDecodeError:
                    csv_text = file_content.decode('iso-8859-1')
                sample = csv_text[:4000] if len(csv_text) > 4000 else csv_text
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except Exception:
                    dialect = csv.excel
                reader = csv.reader(io.StringIO(csv_text), dialect)
                first_row = next(reader, None)
                if not first_row:
                    raise UserError(_("Le fichier CSV est vide."))
                headers = [h.strip() for h in first_row if h]
                for r in reader:
                    if not r or not any(val.strip() for val in r):
                        continue
                    row_dict = {}
                    for idx, val in enumerate(r):
                        if idx < len(headers):
                            row_dict[headers[idx]] = val.strip()
                    rows.append(row_dict)
            except UserError:
                raise
            except Exception as e:
                raise UserError(_("Erreur lors de la lecture du fichier CSV : %s") % str(e))

        if not rows:
            raise UserError(_("Aucune ligne de données trouvée dans le fichier."))
        return headers, rows

    # ------------------------------------------------------------------ #
    # Wizard actions
    # ------------------------------------------------------------------ #
    def _reopen_form(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'campaign.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def action_open_template_wizard(self):
        """Generate and download the XLSX template directly."""
        template_wizard = self.env['contact.template.wizard'].create({})
        return template_wizard.action_download_template()

    def action_analyze_file(self):
        """Step 1 → 2: parse file, detect columns, propose mappings, show review."""
        self.ensure_one()
        try:
            self._update_partner_form_view()
        except Exception as e:
            import logging
            logging.getLogger(__name__).error("Failed to restore dynamic view: %s", e)
        headers, rows = self._parse_uploaded_file()

        gemini_api_key = self.env['ir.config_parameter'].sudo().get_param(
            'ceo_campaign_automation.gemini_api_key', '')
        partner_fields = self.env['res.partner'].fields_get(
            attributes=['string', 'type', 'required'])

        # Drop previous review lines if user re-analyzes
        self.column_line_ids.unlink()

        line_vals = []
        for header in headers:
            samples = [row.get(header) for row in rows[:5] if row.get(header)]
            target_field, label, is_new = self._detect_field_mapping(
                header, samples, partner_fields, gemini_api_key)
            line_vals.append((0, 0, {
                'header': header,
                'sample_preview': " | ".join(samples[:3])[:120],
                'suggested_field': target_field,
                'is_new': bool(is_new),
                # Default behaviour: include the column.
                # For NEW fields, check the create field box by default
                'create_field': bool(is_new),
                'field_label': label,
                'field_description': '',
                'selected': True,
            }))

        self.write({
            'column_line_ids': line_vals,
            'total_rows': len(rows),
            'state': 'review',
        })
        return self._reopen_form()

    def action_back_to_draft(self):
        """Allow the user to upload a different file / restart from scratch."""
        self.ensure_one()
        self.column_line_ids.unlink()
        self.write({'state': 'draft', 'total_rows': 0})
        return self._reopen_form()

    def action_confirm_import(self):
        """Step 2 → 3: apply user choices and create the partner records."""
        self.ensure_one()
        if self.state != 'review':
            raise UserError(_("L'importation n'est pas dans l'état 'Révision'."))

        # Validate that new-field rows the user agreed to create have a label
        missing_labels = self.column_line_ids.filtered(
            lambda l: l.selected and l.is_new and l.create_field and not l.field_label)
        if missing_labels:
            raise UserError(_(
                "Veuillez renseigner un libellé pour les nouveaux champs : %s"
            ) % ", ".join(missing_labels.mapped('header')))

        # Re-parse the file (it is still in self.file_data)
        _headers, rows = self._parse_uploaded_file()

        # Apply row limit
        if self.row_limit and self.row_limit > 0:
            rows = rows[:self.row_limit]

        # Build header → target field map and create the requested new fields
        column_map = {}
        created_fields = []
        skipped_new_fields = []
        for line in self.column_line_ids:
            if not line.selected:
                continue
            if line.is_new:
                if not line.create_field:
                    skipped_new_fields.append(line.header)
                    continue
                fname = self._sanitize_field_name(line.suggested_field or line.header)
                self._create_custom_field(fname, line.field_label, line.field_description)
                created_fields.append(fname)
                column_map[line.header] = fname
            else:
                column_map[line.header] = line.suggested_field

        # Create / update partners (upsert by email then name)
        Partner = self.env['res.partner']
        partner_fields = Partner.fields_get(attributes=['type', 'relation', 'selection'])

        created = 0
        updated = 0
        ignored = 0
        errors = []
        for row in rows:
            vals = {}
            notes = []
            for header, value in row.items():
                if header in skipped_new_fields and value not in (None, ''):
                    notes.append(f"{header}: {value}")

                fname = column_map.get(header)
                if not fname or value in (None, ''):
                    continue

                # Handle relational and boolean fields
                ftype = partner_fields.get(fname, {}).get('type')
                if ftype == 'many2one':
                    relation = partner_fields[fname].get('relation')
                    if relation:
                        record = self.env[relation].search(
                            [('name', 'ilike', str(value).strip())], limit=1)
                        if record:
                            vals[fname] = record.id
                        else:
                            continue  # Skip invalid many2one instead of crashing
                elif ftype == 'many2many':
                    relation = partner_fields[fname].get('relation')
                    if relation:
                        tag_names = [t.strip() for t in str(value).split(',')]
                        tag_ids = []
                        for tn in tag_names:
                            if not tn:
                                continue
                            tag = self.env[relation].search([('name', 'ilike', tn)], limit=1)
                            if not tag:
                                try:
                                    tag = self.env[relation].create({'name': tn})
                                except Exception:
                                    pass
                            if tag:
                                tag_ids.append(tag.id)
                        vals[fname] = [(6, 0, tag_ids)]
                elif ftype == 'boolean':
                    vals[fname] = str(value).strip().lower() in (
                        'true', '1', 'oui', 'yes', 'vrai')
                elif ftype == 'selection':
                    val_lower = str(value).strip().lower()
                    matched_key = value
                    hardcoded_map = {
                        'type': {
                            'contact': 'contact', 'invoice address': 'invoice',
                            'delivery address': 'delivery', 'other address': 'other',
                            'private address': 'private'
                        },
                        'company_type': {
                            'company': 'company', 'person': 'person'
                        }
                    }
                    if fname in hardcoded_map and val_lower in hardcoded_map[fname]:
                        matched_key = hardcoded_map[fname][val_lower]
                    else:
                        sel_options = partner_fields[fname].get('selection')
                        if sel_options and isinstance(sel_options, list):
                            for k, s in sel_options:
                                if (str(s).strip().lower() == val_lower
                                        or str(k).strip().lower() == val_lower):
                                    matched_key = k
                                    break
                    vals[fname] = matched_key
                else:
                    vals[fname] = value

            if notes:
                existing_comment = vals.get('comment', '')
                notes_text = "\n".join(notes)
                vals['comment'] = (
                    f"{existing_comment}\n\n---\nInformations Importées:\n{notes_text}"
                    if existing_comment
                    else f"Informations Importées:\n{notes_text}"
                )

            if not vals.get('name'):
                ignored += 1
                continue

            try:
                # ── Upsert: match by email first, then by exact name ──────
                existing = None
                email_val = vals.get('email', '').strip().lower()
                if email_val:
                    existing = Partner.search(
                        [('email', '=ilike', email_val)], limit=1)
                if not existing:
                    existing = Partner.search(
                        [('name', '=', vals.get('name', ''))], limit=1)

                if existing:
                    existing.write(vals)
                    updated += 1
                else:
                    Partner.create(vals)
                    created += 1
            except Exception as e:
                ignored += 1
                if len(errors) < 10:
                    errors.append("%s — %s" % (vals.get('name', '?'), str(e)))

        # Build summary
        summary_lines = [
            _("Source : %s") % dict(self._fields['source_tag'].selection).get(self.source_tag),
            _("Lignes traitées : %d") % len(rows),
            _("Contacts créés : %d") % created,
            _("Contacts mis à jour : %d") % updated,
            _("Lignes ignorées : %d") % ignored,
            _("Colonnes importées : %d / %d") % (
                len(column_map), len(self.column_line_ids)),
        ]
        if created_fields:
            summary_lines.append(
                _("Nouveaux champs créés : %s") % ", ".join(created_fields))
        if skipped_new_fields:
            summary_lines.append(
                _("Colonnes ignorées (création de champ refusée) : %s")
                % ", ".join(skipped_new_fields))
        if errors:
            summary_lines.append("")
            summary_lines.append(_("Erreurs (10 premières) :"))
            summary_lines.extend(errors)

        self.write({'state': 'done', 'summary': "\n".join(summary_lines)})
        return self._reopen_form()


class CampaignImportWizardLine(models.TransientModel):
    _name = 'campaign.import.wizard.line'
    _description = 'Colonne détectée dans le fichier importé'
    _order = 'is_new desc, header'

    wizard_id = fields.Many2one(
        'campaign.import.wizard', required=True, ondelete='cascade')

    header = fields.Char(string="Colonne du fichier", readonly=True)
    sample_preview = fields.Char(string="Aperçu", readonly=True)

    selected = fields.Boolean(
        string="Importer", default=True,
        help="Décochez pour ignorer cette colonne lors de l'importation.")

    suggested_field = fields.Char(string="Champ Odoo", readonly=True)
    is_new = fields.Boolean(
        string="Nouveau", readonly=True,
        help="Si coché, ce champ n'existe pas encore dans Odoo et doit être créé.")

    # Confirmation required for the creation of new custom fields
    create_field = fields.Boolean(
        string="Créer le champ",
        help="Cochez pour autoriser la création de ce nouveau champ personnalisé "
             "dans res.partner.")
    field_label = fields.Char(
        string="Libellé du champ",
        help="Libellé qui sera affiché dans le formulaire du contact.")
    field_description = fields.Char(
        string="Description",
        help="Texte d'aide (tooltip) affiché à l'utilisateur sur ce champ.")
