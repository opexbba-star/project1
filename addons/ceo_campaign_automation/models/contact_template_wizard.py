# -*- coding: utf-8 -*-
import io
import base64
from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# Standard column definitions
# Each dict has:
#   field   – Odoo field technical name (used as column header for import)
#   label   – Human-readable column header (Row 1)
#   help    – Short description shown in Row 2
#   example – Sample value shown in Row 3
#   type    – Field type (char / selection / many2one / boolean …)
#   options – (optional) static list of allowed values → triggers a dropdown
#   dynamic – (optional) 'languages' | 'titles'  → populated at runtime
# ─────────────────────────────────────────────────────────────────────────────
STANDARD_COLUMNS = [
    {
        'field': 'name',
        'label': 'Name',
        'help': 'Contact full name',
        'example': 'John Doe',
        'type': 'char',
    },
    {
        'field': 'company_type',
        'label': 'Company Type',
        'help': 'Company or Person',
        'example': 'Company',
        'type': 'selection',
        'options': ['Company', 'Person'],
    },
    {
        'field': 'parent_id',
        'label': 'Parent Company',
        'help': 'Display name of the parent company (for contacts / addresses)',
        'example': 'ACME Corporation',
        'type': 'many2one',
    },
    {
        'field': 'type',
        'label': 'Address Type',
        'help': 'Contact / Invoice Address / Delivery Address / Other Address',
        'example': 'Contact',
        'type': 'selection',
        'options': ['Contact', 'Invoice Address', 'Delivery Address', 'Other Address'],
    },
    {
        'field': 'email',
        'label': 'Email',
        'help': 'Primary email address',
        'example': 'contact@example.com',
        'type': 'char',
    },
    {
        'field': 'phone',
        'label': 'Phone',
        'help': 'Main phone number (e.g. +1 555 000 0000)',
        'example': '+1 555 000 0000',
        'type': 'char',
    },
    {
        'field': 'mobile',
        'label': 'Mobile',
        'help': 'Mobile / cell phone number',
        'example': '+1 555 000 0001',
        'type': 'char',
    },
    {
        'field': 'website',
        'label': 'Website',
        'help': 'Company or personal website URL',
        'example': 'https://www.example.com',
        'type': 'char',
    },
    {
        'field': 'title',
        'label': 'Title',
        'help': 'Salutation title (e.g. Mr., Ms., Dr.) — loaded from system',
        'example': 'Mr.',
        'type': 'selection',
        'dynamic': 'titles',
    },
    {
        'field': 'function',
        'label': 'Job Position',
        'help': 'Job title or professional function',
        'example': 'Sales Manager',
        'type': 'char',
    },
    {
        'field': 'lang',
        'label': 'Language',
        'help': 'Language code (e.g. en_US) — installed languages loaded from system',
        'example': 'en_US',
        'type': 'selection',
        'dynamic': 'languages',
    },
    {
        'field': 'category_id',
        'label': 'Tags',
        'help': 'Comma-separated tag names (e.g. VIP, Customer) — loaded from system tags',
        'example': 'VIP, Customer',
        'type': 'char',
    },
    {
        'field': 'comment',
        'label': 'Notes',
        'help': 'Internal notes or free-text comments',
        'example': 'Prefers contact via email only',
        'type': 'char',
    },
    {
        'field': 'ref',
        'label': 'Reference',
        'help': 'Internal reference or customer number',
        'example': 'CUST-001',
        'type': 'char',
    },
    {
        'field': 'active',
        'label': 'Active',
        'help': 'True to keep the record active, False to archive it',
        'example': 'True',
        'type': 'selection',
        'options': ['True', 'False'],
    },
    {
        'field': 'vat',
        'label': 'Tax ID',
        'help': 'VAT or Tax Identification Number',
        'example': 'US123456789',
        'type': 'char',
    },
    {
        'field': 'street',
        'label': 'Street',
        'help': 'Street address — line 1',
        'example': '123 Main Street',
        'type': 'char',
    },
    {
        'field': 'street2',
        'label': 'Street2',
        'help': 'Street address — line 2 (apartment, suite, etc.)',
        'example': 'Suite 100',
        'type': 'char',
    },
    {
        'field': 'city',
        'label': 'City',
        'help': 'City or locality name',
        'example': 'New York',
        'type': 'char',
    },
    {
        'field': 'state_id',
        'label': 'State',
        'help': 'State or province — enter the full display name exactly',
        'example': 'New York',
        'type': 'many2one',
    },
    {
        'field': 'zip',
        'label': 'Zip',
        'help': 'Postal / ZIP code',
        'example': '10001',
        'type': 'char',
    },
    {
        'field': 'country_id',
        'label': 'Country',
        'help': 'Country — enter the full country name exactly',
        'example': 'United States',
        'type': 'many2one',
    },
]

# Field names already covered by the standard columns (used to skip duplicates
# when scanning for custom fields).
_STANDARD_FIELD_NAMES = frozenset(col['field'] for col in STANDARD_COLUMNS)

# ttype values that cannot be meaningfully represented in a flat spreadsheet.
_SKIP_TYPES = frozenset(['one2many', 'many2many', 'binary'])


class ContactTemplateWizard(models.TransientModel):
    _name = 'contact.template.wizard'
    _description = 'Download Contacts Import XLSX Template'

    # ── Dummy field so the form view has something to bind ────────────────────
    # (TransientModel with no stored fields is fine in Odoo, but some versions
    #  need at least the _name defined; we add an informational readonly field
    #  so the view can display dynamic tag info.)
    tag_list_preview = fields.Char(
        string='Existing Tags (preview)',
        compute='_compute_tag_list_preview',
    )

    def _compute_tag_list_preview(self):
        tags = self.env['res.partner.category'].search([], limit=10)
        tag_names = ', '.join(tags.mapped('name'))
        suffix = ' …' if len(tags) == 10 else ''
        for rec in self:
            rec.tag_list_preview = tag_names + suffix if tag_names else _('(no tags found)')

    # ── Dynamic option loaders ─────────────────────────────────────────────────

    def _get_language_options(self):
        """Return [(code, name)] for all active installed languages."""
        langs = self.env['res.lang'].search([('active', '=', True)], order='name')
        return [(lang.code, lang.name) for lang in langs]

    def _get_title_options(self):
        """Return a list of title names from res.partner.title."""
        if 'res.partner.title' in self.env:
            titles = self.env['res.partner.title'].search([], order='name')
            return [t.name for t in titles] or ['Mr.', 'Ms.', 'Dr.']
        return ['Mr.', 'Ms.', 'Dr.']

    def _get_tag_options(self):
        """Return a list of all partner tag names."""
        cats = self.env['res.partner.category'].search([], order='name')
        return [c.name for c in cats]

    def _get_custom_fields(self):
        """
        Return a list of column-dicts for manual (x_*) fields on res.partner.
        Excludes:
          - Fields already in STANDARD_COLUMNS
          - Relational / binary types that can't be represented in a flat file
          - Pure computed (non-stored) fields
        """
        records = self.env['ir.model.fields'].search([
            ('model', '=', 'res.partner'),
            ('state', '=', 'manual'),
            ('ttype', 'not in', list(_SKIP_TYPES)),
        ], order='field_description')

        columns = []
        for rec in records:
            if rec.name in _STANDARD_FIELD_NAMES:
                continue
            # Skip non-stored computed fields
            if rec.compute and not rec.store:
                continue

            col = {
                'field': rec.name,
                'label': rec.field_description or rec.name,
                'help': rec.help or ('Custom field: ' + (rec.field_description or rec.name)),
                'example': '',
                'type': rec.ttype,
            }

            if rec.ttype == 'selection':
                # Fetch the human-readable option labels
                info = self.env['res.partner'].fields_get(
                    [rec.name], attributes=['selection'])
                sel = info.get(rec.name, {}).get('selection', [])
                if sel:
                    col['options'] = [label for _val, label in sel]

            elif rec.ttype == 'boolean':
                col['options'] = ['True', 'False']
                col['example'] = 'True'

            columns.append(col)

        return columns

    # ── XLSX builder ───────────────────────────────────────────────────────────

    def _build_column_list(self):
        """Merge standard columns with runtime-resolved dynamic options and custom fields."""
        lang_options   = self._get_language_options()   # [(code, name), …]
        title_options  = self._get_title_options()       # ['Mr.', …]

        all_columns = []
        for col in STANDARD_COLUMNS:
            c = dict(col)  # shallow copy — don't mutate the module-level list

            if c.get('dynamic') == 'languages':
                codes = [code for code, _name in lang_options]
                c['options'] = codes
                preview = ', '.join(
                    f"{name} ({code})" for code, name in lang_options[:5]
                )
                if len(lang_options) > 5:
                    preview += ' …'
                c['help'] = 'Language code — ' + preview if preview else c['help']

            elif c.get('dynamic') == 'titles':
                c['options'] = title_options

            all_columns.append(c)

        # Append custom fields after the 22 standard ones
        all_columns.extend(self._get_custom_fields())
        return all_columns

    def _generate_xlsx(self):
        """Build the XLSX workbook and return its raw bytes."""
        if not OPENPYXL_AVAILABLE:
            raise UserError(_(
                "The Python package 'openpyxl' is not installed on this server.\n"
                "Please ask your system administrator to install it "
                "(pip install openpyxl)."
            ))

        all_columns = self._build_column_list()

        # ── Workbook & sheets ─────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Contacts Import'
        ws.sheet_properties.tabColor = '1F3864'

        # Hidden helper sheet that stores dropdown option lists
        # (avoids the 255-char limit on inline DataValidation formula strings)
        lists_ws = wb.create_sheet('_lists')
        lists_ws.sheet_state = 'hidden'

        # ── Styles ────────────────────────────────────────────────────────────
        thin_side   = Side(style='thin', color='BDD7EE')
        cell_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side,  bottom=thin_side,
        )

        header_font  = Font(name='Calibri', bold=True,   color='FFFFFF', size=11)
        header_fill  = PatternFill(fill_type='solid', fgColor='1F3864')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        help_font    = Font(name='Calibri', italic=True, color='1A3A5C', size=10)
        help_fill    = PatternFill(fill_type='solid', fgColor='D6E4F0')
        help_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        ex_font      = Font(name='Calibri', size=10, color='333333')
        ex_fill      = PatternFill(fill_type='solid', fgColor='FFFFFF')
        ex_align     = Alignment(horizontal='left',   vertical='center', wrap_text=False)

        # ── Write Rows 1–3 and register dropdowns ─────────────────────────────
        for col_idx, col in enumerate(all_columns, start=1):
            col_letter = get_column_letter(col_idx)

            # Row 1 – Header label
            cell = ws.cell(row=1, column=col_idx, value=col['label'])
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = cell_border

            # Row 2 – Help / description
            cell = ws.cell(row=2, column=col_idx, value=col['help'])
            cell.font      = help_font
            cell.fill      = help_fill
            cell.alignment = help_align
            cell.border    = cell_border

            # Row 3 – Example value
            cell = ws.cell(row=3, column=col_idx, value=col['example'])
            cell.font      = ex_font
            cell.fill      = ex_fill
            cell.alignment = ex_align
            cell.border    = cell_border

            # Dropdown validation (selection / boolean fields with options list)
            options = col.get('options')
            if options:
                self._add_dropdown(ws, lists_ws, col_letter, col_idx, options)

        # ── Freeze top 3 rows ─────────────────────────────────────────────────
        ws.freeze_panes = 'A4'

        # ── Row heights ───────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 30   # header
        ws.row_dimensions[2].height = 48   # help (may wrap)
        ws.row_dimensions[3].height = 20   # example

        # ── Column widths (auto-sized, clamped 15–55 chars) ───────────────────
        for col_idx, col in enumerate(all_columns, start=1):
            max_len = max(
                len(col.get('label', '')),
                min(len(col.get('help', '')), 45),
                len(col.get('example', '')),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(15, min(max_len + 4, 55))

        # ── Output ────────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Dropdown helper ───────────────────────────────────────────────────────

    def _add_dropdown(self, ws, lists_ws, col_letter, col_idx, options):
        """
        Write *options* into the hidden _lists sheet (one column per main column)
        and attach a DataValidation list dropdown to the data rows (4–1004) of
        the given *col_letter* in the main sheet.
        """
        if not options:
            return

        # Write options into _lists sheet (same column index as the main sheet)
        list_col_letter = get_column_letter(col_idx)
        for row_i, opt in enumerate(options, start=1):
            lists_ws.cell(row=row_i, column=col_idx, value=str(opt))

        # Build range reference
        list_range = f"'_lists'!${list_col_letter}$1:${list_col_letter}${len(options)}"

        dv = DataValidation(
            type='list',
            formula1=list_range,
            allow_blank=True,
            # Note: showDropDown=True HIDES the dropdown arrow in Excel (inverted logic).
            # We intentionally leave it unset (None) so the arrow IS visible.
        )
        dv.error       = 'Please choose a value from the dropdown list.'
        dv.errorTitle  = 'Invalid entry'
        dv.prompt      = 'Select a value from the list'
        dv.promptTitle = 'Allowed values'

        ws.add_data_validation(dv)
        dv.add(f'{col_letter}4:{col_letter}1004')

    # ── Public action ──────────────────────────────────────────────────────────

    def action_download_template(self):
        """Generate the XLSX template and return a file download action."""
        self.ensure_one()

        xlsx_bytes = self._generate_xlsx()

        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'contacts_import_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(xlsx_bytes).decode('utf-8'),
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
